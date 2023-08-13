from openpyxl import load_workbook
from BallerClass import Baller
from BallerClass import Batter

global first_name, last_name, age, runs, balls_played, is_on_strike, is_batting,\
    wickets, runs_given, is_balling, no_of_balls
pak_bat = []
pak_ball = []

# Load the workbook
wb = load_workbook('Pakistan CT.xlsx')
# accessing first sheet by its name 'Bat'
ws = wb['Bat']
for r in range(2, 13):  # 2 to 12+1
    first_name = str(ws.cell(row=r, column=1).value)
    last_name = str(ws.cell(row=r, column=2).value)
    age = int(ws.cell(row=r, column=3).value)
    runs = int(ws.cell(row=r, column=4).value)
    balls_played = int(ws.cell(row=r, column=5).value)
    is_on_strike = bool(ws.cell(row=r, column=6).value)
    is_batting = bool(ws.cell(row=r, column=7).value)
    pak_bat.append(Batter(first_name, last_name, age, runs, balls_played, is_on_strike, is_batting))

# accessing first sheet by its name 'Ball'
ws = wb['Ball']
for r in range(2, 7):  # 2 to 6+1
    first_name = str(ws.cell(row=r, column=1).value)
    last_name = str(ws.cell(row=r, column=2).value)
    age = int(ws.cell(row=r, column=3).value)
    runs = int(ws.cell(row=r, column=4).value)
    balls_played = int(ws.cell(row=r, column=5).value)
    is_on_strike = bool(ws.cell(row=r, column=6).value)
    is_batting = bool(ws.cell(row=r, column=7).value)
    wickets = int(ws.cell(row=r, column=8).value)
    runs_given = int(ws.cell(row=r, column=9).value)
    is_balling = bool(ws.cell(row=r, column=10).value)
    no_of_balls = int(ws.cell(row=r, column=11).value)
    pak_ball.append(Baller(first_name, last_name, age, runs, balls_played, is_on_strike, is_batting,
                           wickets, runs_given, is_balling, no_of_balls))
