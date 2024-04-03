from BowlerClass import Bowler
from BatterClass import Batter
from openpyxl import load_workbook
import functions as f


class Team:
    """Represents a cricket team with its players (batters and bowlers)."""

    def __init__(self, name):
        """Initializes a Team object with its name and empty player lists."""
        self.name = name
        self.batters = []  # Initialize array for batters
        self.bowlers = []  # Initialize array for bowlers
        self.current_ball = 0
        self.current_over = 0.0
        self.current_batters = None
        self.current_bowler = None
        self.total_team_score = 0  # Total team score
        self.total_team_wickets = 0  # Total team Wickets
        self.total_extras = 0  # Total extras
        self.run_outs = 0
        self.current_runrate = 0

        # Add players during initialization
        self.add_players(self.name)

    def add_players(self, name):
        """Loads player data from an Excel spreadsheet and creates Batter and Bowler objects."""
        print("Adding players")
        try:
            # Load the workbook
            wb = load_workbook('Teams.xlsx')  # Open the Teams.xlsx file

            # Access the sheet by its name
            ws = wb[name]  # Select the sheet with the specified name

            # Add batters
            for r in range(2, 13):  # Iterate through rows 2 to 12 for batters
                data = [str(ws.cell(row=r, column=i).value) for i in range(1, 4)]  # Extract data from cells
                first_name, last_name, age = data  # Unpack player data
                batter = Batter(  # Create a Batter object
                    first_name,
                    last_name,
                    int(age)
                )
                self.batters.append(batter)  # Add the batter to the team

            self.current_batters = [self.batters[0], self.batters[1]]  # Current 2 batters on pitch

            # Add bowlers (using the same sheet, indices 8 to 12)
            for r in range(8, 13):  # Iterate through rows 8 to 12 for bowlers
                data = [str(ws.cell(row=r, column=i).value) for i in range(1, 4)]  # Extract data from cells
                first_name, last_name, age = data  # Unpack player data
                bowler = Bowler(  # Create a Bowler object
                    first_name,
                    last_name,
                    int(age)
                )
                self.bowlers.append(bowler)  # Add the bowler to the team

            self.current_bowler = self.bowlers[0]  # Current bowler

        except FileNotFoundError:
            print("Error: Teams.xlsx file not found.")
        except KeyError:
            print(f"Error: Sheet '{name}' not found in the workbook.")

    # Setter Functions --------------------------------
    def set_over(self):
        self.current_over = round(float(int(self.current_ball / 6) + (int(self.current_ball % 6)) * 0.1), 1)

    def set_current_bowler(self, total_overs):
        if total_overs > self.get_int_over() == self.current_over:
            self.current_bowler = self.bowlers[f.display_choices(choices=self.bowlers, choice_type="Bowler",
                                                                 is_bowler=True)]

    def set_current_runrate(self):
        if self.current_ball > 0:
            # runs per over
            self.current_runrate = (self.total_team_score * 6) / self.current_ball
        else:
            self.current_runrate = 0

    # Increment Functions --------------------------------
    def increment_total_extras(self, total_extras):
        self.total_extras += total_extras

    def increment_wickets(self):
        self.total_team_wickets += 1

    def increment_runs_t(self, runs_scored):
        """Increments the team's total runs"""
        self.total_team_score += runs_scored

        # Increment ball and over count
        self.increment_ball()

    def increment_team_score(self, runs):
        self.total_team_score += runs

    def increment_ball(self):
        self.current_ball += 1
        self.set_over()

    def increment_run_outs(self):
        self.run_outs += 1

    # Getter Functions --------------------------------

    def get_total_team_score(self):
        return self.total_team_score

    def get_total_wickets(self):
        return self.total_team_wickets

    def get_int_over(self):
        return int(self.current_over)

    def get_current_bowler(self, total_overs):
        # done_flag = 0
        # self.set_current_bowler(total_overs=total_overs)
        return self.current_bowler

    def get_current_runrate(self):
        self.set_current_runrate()
        return self.current_runrate
