import openpyxl
from flask import Flask, render_template, request
from TeamClass import Team
from MatchClass import Match

app = Flask(__name__)


@app.route('/')
def index():
    # Assuming Teams.xlsx contains team names in individual sheets
    team_names = sorted([sheet.title for sheet in openpyxl.load_workbook('D:\My Code\Cricket\Teams.xlsx').worksheets])
    return render_template('index.html', teams=team_names)


@app.route('/simulate', methods=['POST'])
def simulate():
    try:
        team_A_name = request.form['teamA']
        team_B_name = request.form['teamB']
        match_type = request.form['matchType']
        sheet_name = request.form['stadiumSheet']  # Assuming stadium names are in a sheet

        # Load stadium names from sheet
        workbook = openpyxl.load_workbook('Locations.xlsx')
        worksheet = workbook[sheet_name]
        stadiums = []
        for row in worksheet.iter_rows(values_only=True, min_row=2):
            for cell in row:
                if cell:
                    stadiums.append(cell)
        stadium_name = request.form['stadium']

        team_A = Team(team_A_name)
        team_B = Team(team_B_name)
        match = Match(team1=team_A, team2=team_B, match_type=match_type, stadium=stadium_name)
        match.simulate_match()  # Assuming this method generates results

        # Replace with your actual logic for fetching results
        winner = "Team A"  # Replace with actual winner
        scorecard = {"Team A": 150, "Team B": 145}  # Replace with actual scorecard

        return render_template('results.html', winner=winner, scorecard=scorecard)

    except FileNotFoundError as e:
        return render_template('error.html', error_message=str(e))


if __name__ == '__main__':
    app.run(debug=True)
