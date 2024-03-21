import openpyxl
from TeamClass import Team
from MatchClass import Match


def main():
    """Loads teams from an Excel file, prompts for team selection,
       creates Team objects, simulates a match, and displays results."""

    team_A, team_B = choose_team()
    match_type = choose_match_type()
    stadium_name = choose_location()

    match = Match(team1=team_A, team2=team_B, match_type=match_type, stadium=stadium_name)
    match.simulate_match()


def choose_location():
    try:
        workbook = openpyxl.load_workbook('Locations.xlsx')
    except FileNotFoundError:
        print("Error: Locations.xlsx file not found.")
        return

    # Extracting sheet names from the workbook
    sheet_names = workbook.sheetnames

    # Display available sheets to choose from
    display_choices(sheet_names, "Choose a sheet for locations")

    # Prompt user to select a sheet
    sheet_name = get_choice(sheet_names, "sheet")
    worksheet = workbook[sheet_name]
    stadiums = []
    for row in worksheet.iter_rows(values_only=True, min_row=2):
        for cell in row:
            if cell:
                stadiums.append(cell)
    print(stadiums)
    display_choices(choices=stadiums,choice_type='Choose a Stadium')
    stadium_name = get_choice(choices=stadiums,choice_type='Stadium')
    return stadium_name


def choose_team():
    try:
        workbook = openpyxl.load_workbook('Teams.xlsx')
    except FileNotFoundError:
        print("Error: Teams.xlsx file not found.")
        return

    team_names = sorted([sheet.title for sheet in workbook.worksheets])

    display_choices(choices=team_names, choice_type="Choose Team A")
    team_A_name = get_choice(choices=team_names, choice_type="Team A")
    team_names.remove(team_A_name)

    display_choices(choices=team_names, choice_type="Choose Team B")
    team_B_name = get_choice(choices=team_names, choice_type="Team B")

    team_A = Team(team_A_name)
    team_B = Team(team_B_name)
    return team_A, team_B


def choose_match_type():
    match_types = ["Super Over: 1 Over Match", "Fives: 5 Overs Match", "T10: 10 Overs Match",
                   "T20: 20 Overs Match", "ODI: 50 Overs Match"]
    display_choices(choices=match_types, choice_type="Choose Match Type")
    return get_choice(choices=match_types, choice_type="Match Type")


def display_choices(choices, choice_type):
    """Displays numbered options for team selection."""
    print(f'_______{choice_type}_______')
    for i, choice in enumerate(choices, start=1):
        print(f"{i}: {choice}")
    print()


def get_choice(choices, choice_type):
    """Prompts for team selection and validates input."""

    while True:
        choice = input(f"Enter {choice_type} (1-{len(choices)}): ")
        try:
            choice = int(choice)
            if 1 <= choice <= len(choices):
                return choices[choice - 1]
            else:
                print("Invalid choice. Please enter a number between 1 and", len(choices))
        except ValueError:
            print("Invalid input. Please enter a number.")


if __name__ == "__main__":
    main()

# TO-DO
# for Match > 5 over it bowler repeats
